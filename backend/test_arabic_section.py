import asyncio
import io
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

import server
from server import (
    SCHOOL_SECTION_ARABIC,
    SCHOOL_SECTION_INTERNATIONAL,
    arabic_educational_stage_for_grade,
    arabic_exam_raw_max_for_grade,
    arabic_score_summary,
    build_arabic_grading_payload,
    ClassBase,
    ClassUpdate,
    clear_arabic_class_grades,
    classify_legacy_arabic_practical,
    create_class,
    generate_arabic_grades_excel,
    generate_arabic_grades_pdf,
    school_section_query,
    parse_class_name,
    migrate_arabic_class_metadata,
    update_class,
    validate_arabic_exam_values,
)
from test_arabic_import import _matches


class _Cursor:
    def __init__(self, rows):
        self.rows = list(rows)

    def sort(self, *_args):
        return self

    async def to_list(self, limit):
        return [dict(row) for row in self.rows[:limit]]

    def __aiter__(self):
        self._index = 0
        return self

    async def __anext__(self):
        if self._index >= len(self.rows):
            raise StopAsyncIteration
        row = dict(self.rows[self._index])
        self._index += 1
        return row


class _Collection:
    def __init__(self, rows=()):
        self.rows = [dict(row) for row in rows]

    def find(self, query=None, projection=None):
        rows = [row for row in self.rows if _matches(row, query or {})]
        if projection:
            included = [key for key, value in projection.items() if value and key != "_id"]
            if included:
                rows = [{key: row[key] for key in included if key in row} for row in rows]
            elif projection.get("_id") == 0:
                rows = [{key: value for key, value in row.items() if key != "_id"} for row in rows]
        return _Cursor(rows)

    async def update_one(self, query, update):
        for row in self.rows:
            if _matches(row, query):
                row.update(update.get("$set", {}))
                return SimpleNamespace(matched_count=1)
        return SimpleNamespace(matched_count=0)

    async def find_one(self, query, projection=None):
        for row in self.rows:
            if _matches(row, query):
                result = dict(row)
                if projection and projection.get("_id") == 0:
                    result.pop("_id", None)
                return result
        return None

    async def find_one_and_update(self, query, update, return_document=None):
        for row in self.rows:
            if _matches(row, query):
                row.update(update.get("$set", {}))
                return dict(row)
        return None

    async def delete_many(self, query):
        before = len(self.rows)
        self.rows = [row for row in self.rows if not _matches(row, query)]
        return SimpleNamespace(deleted_count=before - len(self.rows))


def test_primary_uses_best_theory_and_weights_each_contribution_to_30():
    result = arabic_score_summary(
        {"theory_test_1": 11, "theory_test_2": 13, "practical_test": 12}, exam_raw_max=15
    )
    assert result["best_theory_raw"] == 13
    assert result["best_theory_weighted"] == 26
    assert result["practical_weighted"] == 24
    assert result["tests_total"] == 50


def test_middle_secondary_uses_raw_max_20_without_early_rounding():
    result = arabic_score_summary(
        {"theory_test_1": 14, "theory_test_2": 17, "practical_test": 18}, exam_raw_max=20
    )
    assert result["best_theory_raw"] == 17
    assert result["best_theory_weighted"] == 25.5
    assert result["practical_weighted"] == 27
    assert result["tests_total"] == 52.5


@pytest.mark.parametrize(
    ("theory_1", "theory_2", "expected"),
    [(10, 13, 13), (15, 12, 15), (0, 8, 8), (0, None, 0), (None, 11, 11), (None, None, None)],
)
def test_best_theory_preserves_zero_and_null(theory_1, theory_2, expected):
    result = arabic_score_summary(
        {"theory_test_1": theory_1, "theory_test_2": theory_2}, exam_raw_max=15
    )
    assert result["best_theory_raw"] == expected


def test_completion_requires_all_three_attempts_but_allows_provisional_total():
    result = arabic_score_summary(
        {"performance_tasks": 10, "theory_test_1": 0, "theory_test_2": None, "practical_test": None},
        exam_raw_max=15,
    )
    assert result["quarter_total"] == 10
    assert result["test_completion"] == {
        "theory_test_1": True,
        "theory_test_2": False,
        "practical_test": False,
    }
    assert result["test_completion_count"] == 1
    assert result["all_tests_completed"] is False

    complete_zero = arabic_score_summary(
        {"theory_test_1": 0, "theory_test_2": 0, "practical_test": 0}, exam_raw_max=15
    )
    assert complete_zero["all_tests_completed"] is True
    assert complete_zero["tests_total"] == 0
    assert complete_zero["quarter_total"] == 0


def test_arabic_blank_quarter_stays_no_data():
    result = arabic_score_summary({})
    assert result["quarter_total"] is None
    assert result["has_grades"] is False
    assert result["test_completion_count"] == 0


def test_clear_arabic_grades_is_limited_to_one_class_and_exact_term():
    fake_db = SimpleNamespace(
        classes=_Collection([
            {"id": "c4a", "name": "رابع أ", "school_section": "arabic", "academic_year": "2026-2027"},
            {"id": "c4b", "name": "رابع ب", "school_section": "arabic", "academic_year": "2026-2027"},
        ]),
        students=_Collection([
            {"id": "s1", "class_id": "c4a", "school_section": "arabic", "academic_year": "2026-2027"},
            {"id": "s2", "class_id": "c4b", "school_section": "arabic", "academic_year": "2026-2027"},
        ]),
        arabic_quarter_scores=_Collection([
            {"id": "delete", "student_id": "s1", "academic_year": "2026-2027", "semester": 1, "quarter": 1},
            {"id": "keep-term", "student_id": "s1", "academic_year": "2026-2027", "semester": 1, "quarter": 2},
            {"id": "keep-class", "student_id": "s2", "academic_year": "2026-2027", "semester": 1, "quarter": 1},
        ]),
    )
    teacher = {"id": "t1", "role_name": "Teacher", "assigned_class_ids": ["c4a"]}
    with patch.object(server, "db", fake_db), patch.object(server, "log_user_action", AsyncMock()):
        result = asyncio.run(clear_arabic_class_grades("2026-2027", 1, 1, "c4a", teacher))
    assert result["grades_deleted"] == 1
    assert result["students_in_scope"] == 1
    assert {row["id"] for row in fake_db.arabic_quarter_scores.rows} == {"keep-term", "keep-class"}

    with patch.object(server, "db", fake_db), patch.object(server, "log_user_action", AsyncMock()), pytest.raises(HTTPException) as error:
        asyncio.run(clear_arabic_class_grades("2026-2027", 1, 1, "c4b", teacher))
    assert error.value.status_code == 403
    assert error.value.detail == "arabic_grade_clear_forbidden"


def test_stage_is_derived_from_canonical_class_grade():
    assert arabic_educational_stage_for_grade(4) == "primary"
    assert arabic_exam_raw_max_for_grade(6) == 15
    assert arabic_educational_stage_for_grade(7) == "middle"
    assert arabic_exam_raw_max_for_grade(9) == 20
    assert arabic_educational_stage_for_grade(10) == "secondary"
    assert arabic_exam_raw_max_for_grade(12) == 20
    with pytest.raises(ValueError):
        arabic_exam_raw_max_for_grade(None)


@pytest.mark.parametrize(
    ("name", "grade", "section"),
    [
        ("رابع أ", 4, "A"), ("رابع ب", 4, "B"), ("خامس أ", 5, "A"),
        ("سادس ب", 6, "B"), ("الصف الحادي عشر أ", 11, "A"), ("٤A", 4, "A"),
    ],
)
def test_arabic_class_names_resolve_canonical_grade_and_section(name, grade, section):
    assert parse_class_name(name) == {"grade": grade, "section": section}


def test_empty_arabic_roster_loads_even_when_legacy_class_name_is_unresolved():
    fake_db = SimpleNamespace(
        classes=_Collection([{
            "_id": "mongo-c1", "id": "c1", "name": "حلقة الحاسب", "grade": None,
            "section": None, "school_section": "arabic", "academic_year": "2026-2027",
        }]),
        students=_Collection(),
        arabic_quarter_scores=_Collection(),
    )
    with patch.object(server, "db", fake_db):
        payload = asyncio.run(build_arabic_grading_payload(
            "2026-2027", 1, 1, None, {"id": "admin", "role_name": "Admin"},
        ))
    assert payload["total_students"] == 0
    assert payload["class_breakdown"][0]["student_count"] == 0
    assert payload["configuration_issues"] == [{
        "code": "arabic_class_grade_required", "class_id": "c1", "class_name": "حلقة الحاسب",
    }]


def test_arabic_class_metadata_migration_repairs_only_unambiguous_legacy_names():
    classes = _Collection([
        {"_id": "m1", "id": "c1", "name": "رابع أ", "grade": None, "section": "رابع أ", "school_section": "arabic"},
        {"_id": "m2", "id": "c2", "name": "حلقة الحاسب", "grade": None, "section": None, "school_section": "arabic"},
    ])
    with patch.object(server, "db", SimpleNamespace(classes=classes)):
        result = asyncio.run(migrate_arabic_class_metadata())
    assert result == {"examined": 2, "updated": 1, "unresolved": 1}
    assert classes.rows[0]["grade"] == 4 and classes.rows[0]["section"] == "A"
    assert classes.rows[1]["grade"] is None


def test_new_arabic_class_requires_resolvable_positive_grade():
    with patch.object(server, "db", SimpleNamespace(classes=_Collection())), pytest.raises(HTTPException) as error:
        asyncio.run(create_class(
            ClassBase(name="حلقة الحاسب", school_section="arabic", academic_year="2026-2027"),
            {"id": "admin", "role_name": "Admin"},
        ))
    assert error.value.status_code == 422
    assert error.value.detail == "arabic_class_grade_required"


def test_renaming_arabic_class_to_unparseable_name_preserves_existing_grade():
    classes = _Collection([{
        "id": "c1", "name": "رابع أ", "grade": 4, "section": "A",
        "school_section": "arabic", "academic_year": "2026-2027",
    }])
    with patch.object(server, "db", SimpleNamespace(classes=classes)):
        result = asyncio.run(update_class("c1", ClassUpdate(name="حلقة الحاسب")))
    assert result["name"] == "حلقة الحاسب"
    assert result["grade"] == 4


def test_unresolved_arabic_class_with_students_returns_stable_configuration_error():
    fake_db = SimpleNamespace(
        classes=_Collection([{
            "id": "c1", "name": "حلقة الحاسب", "grade": None, "section": None,
            "school_section": "arabic", "academic_year": "2026-2027",
        }]),
        students=_Collection([{
            "id": "s1", "full_name": "طالب تجريبي", "class_id": "c1", "class_name": "حلقة الحاسب",
            "school_section": "arabic", "academic_year": "2026-2027",
        }]),
        arabic_quarter_scores=_Collection(),
    )
    with patch.object(server, "db", fake_db), pytest.raises(HTTPException) as error:
        asyncio.run(build_arabic_grading_payload(
            "2026-2027", 1, 1, None, {"id": "admin", "role_name": "Admin"},
        ))
    assert error.value.status_code == 409
    assert error.value.detail == "arabic_class_grade_required"


def test_stage_specific_raw_validation_rejects_scores_above_maximum():
    validate_arabic_exam_values({"theory_test_1": 15, "practical_test": 0}, 15)
    validate_arabic_exam_values({"theory_test_2": 20, "practical_test": 20}, 20)
    with pytest.raises(ValueError):
        validate_arabic_exam_values({"theory_test_1": 15.5}, 15)


def test_legacy_practical_migration_only_maps_unambiguous_values():
    assert classify_legacy_arabic_practical({"practical_test_1": 12}) == {
        "status": "migrated_unambiguous", "practical_test": 12
    }
    assert classify_legacy_arabic_practical({"practical_test_1": 8, "practical_test_2": 8}) == {
        "status": "migrated_unambiguous", "practical_test": 8
    }
    assert classify_legacy_arabic_practical({"practical_test_1": 8, "practical_test_2": 12}) == {
        "status": "manual_review", "practical_test": None
    }
    assert classify_legacy_arabic_practical({
        "practical_test_1": 8, "practical_test_2": 12, "practical_test": 10,
        "legacy_migration_status": "manual_review",
    }) == {"status": "resolved_manual_review", "practical_test": 10}


def test_section_and_academic_year_filters_keep_scopes_isolated():
    international = school_section_query(SCHOOL_SECTION_INTERNATIONAL, "2026-2027")
    arabic_2026 = school_section_query(SCHOOL_SECTION_ARABIC, "2026-2027")
    arabic_2027 = school_section_query(SCHOOL_SECTION_ARABIC, "2027-2028")
    assert {"school_section": {"$exists": False}} in international["$and"][0]["$or"]
    assert arabic_2026 != arabic_2027
    assert arabic_2026 == {
        "$and": [
            {"school_section": SCHOOL_SECTION_ARABIC},
            {"$or": [{"academic_year": "2026-2027"}]},
        ]
    }


def test_quarters_remain_independent_inputs():
    q1 = arabic_score_summary({"theory_test_1": 15, "practical_test": 15}, 15)
    q2 = arabic_score_summary({"theory_test_1": 5, "practical_test": 5}, 15)
    assert q1["tests_total"] == 60
    assert q2["tests_total"] == 20


def test_arabic_pdf_and_excel_exports_render_new_model_rows():
    payload = {
        "academic_year": "2026-2027", "semester": 1, "quarter": 1, "display_quarter": 1,
        "tests_completed": 3, "tests_missing": 0,
        "students": [{
            "full_name": "محمد أحمد", "class_name": "رابع أ", "has_grades": True,
            "exam_raw_max": 15, "theory_test_1": 11, "theory_test_2": 13, "practical_test": 12,
            "continuous_total": 40, "best_theory_weighted": 26, "practical_weighted": 24,
            "tests_total": 50, "quarter_total": 90, "test_completion_count": 3,
        }],
    }
    pdf = generate_arabic_grades_pdf(payload, "ar")
    pdf_en = generate_arabic_grades_pdf(payload, "en")
    excel = generate_arabic_grades_excel(payload, "ar")
    assert pdf.startswith(b"%PDF") and len(pdf) > 2000
    assert pdf_en.startswith(b"%PDF") and len(pdf_en) > 2000
    assert excel.startswith(b"PK") and len(excel) > 2000
    headers = [cell.value for cell in next(load_workbook(io.BytesIO(excel)).active.iter_rows())]
    assert "الاختبار العملي (خام)" in headers
    assert all("العملي 1" not in str(header) and "العملي 2" not in str(header) for header in headers)
