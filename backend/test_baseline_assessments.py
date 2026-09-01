"""Offline API and PDF checks; fixtures contain synthetic students only."""
import copy
import io
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

import server
from baseline_assessments import build_snapshot, level, make_router, percentage
from baseline_pdf import render_baseline_pdf, shaped
from test_arabic_import import _matches
from score_sheet import REFERENCE_HEADERS


class Cursor:
    def __init__(self, rows):
        self.rows = rows

    def sort(self, key, direction):
        self.rows.sort(key=lambda r: str(r.get(key, "")), reverse=direction < 0)
        return self

    async def to_list(self, limit):
        return copy.deepcopy(self.rows[:limit])


class Collection:
    def __init__(self, rows=()):
        self.rows = copy.deepcopy(list(rows))
        self.force_conflict = False

    def find(self, query, projection=None):
        rows = [r for r in self.rows if _matches(r, query)]
        if projection:
            include = [key for key, value in projection.items() if value]
            rows = [{k: v for k, v in r.items() if k in include} if include else {k: v for k, v in r.items() if k not in projection} for r in rows]
        return Cursor(rows)

    async def find_one(self, query):
        return next((copy.deepcopy(r) for r in self.rows if _matches(r, query)), None)

    async def insert_one(self, doc):
        self.rows.append(copy.deepcopy(doc))

    async def update_one(self, query, update):
        for row in self.rows:
            if _matches(row, query) and not self.force_conflict:
                row.update(copy.deepcopy(update.get("$set", {})))
                for key, value in update.get("$inc", {}).items():
                    row[key] += value
                return SimpleNamespace(matched_count=1)
        return SimpleNamespace(matched_count=0)

    async def delete_one(self, query):
        for index, row in enumerate(self.rows):
            if _matches(row, query):
                self.rows.pop(index)
                return SimpleNamespace(deleted_count=1)
        return SimpleNamespace(deleted_count=0)


def sample_record(count=8):
    values = [15, 18, 10, 6, 13, 20, None, 0]
    return {"_id": "sample", "id": "sample", "title": "اختبار قبلي - سجل تجريبي", "test_date": "2026-08-28",
            "teacher_id": "t1", "teacher_name": "معلم تجريبي", "school_section": "international",
            "academic_year": "2026-2027", "semester": 2, "quarter": 1, "max_score": 20.0,
            "revision": 1, "created_at": "2026-08-28", "class_ids": ["c1", "c2"],
            "classes": [{"id": "c1", "name": "4A"}, {"id": "c2", "name": "4B"}],
            "roster": [{"id": f"s{i}", "full_name": f"طالب تجريبي {i+1:02}", "class_id": "c1" if i % 8 < 4 else "c2", "class_name": "4A" if i % 8 < 4 else "4B"} for i in range(count)],
            "scores": {f"s{i}": values[i % 8] for i in range(count)}}


def score_file(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(REFERENCE_HEADERS)
    for name, score in rows:
        sheet.append([name, "لا يستورد", "4A", score, 0.5, "لا يستورد", "لا يستورد", 45, "لا يستورد", "لا يستورد", "لا يستورد"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture
def api_fixture():
    record = sample_record()
    classes = [{**c, "school_section": "international", "academic_year": "2026-2027"} for c in record["classes"]]
    classes += [{"id": "arabic", "name": "4A", "school_section": "arabic", "academic_year": "2026-2027"}]
    students = [{**s, "school_section": "international", "academic_year": "2026-2027"} for s in record["roster"]]
    students += [{"id": "ar-student", "full_name": "طالب عربي", "class_id": "arabic", "school_section": "arabic", "academic_year": "2026-2027"}]
    db = SimpleNamespace(baseline_assessments=Collection([record]), classes=Collection(classes), students=Collection(students))
    state = {"user": {"id": "t1", "role_name": "Teacher", "full_name": "معلم", "assigned_class_ids": ["c1", "c2"]}}

    async def auth():
        return state["user"]

    app = FastAPI()
    app.include_router(make_router(db, auth, server.school_section_query), prefix="/api")
    return TestClient(app), db, state


@pytest.mark.parametrize("maximum", [10, 15, 20, 30, 40])
def test_normalizes_variable_maxima(maximum):
    assert percentage(maximum * .75, maximum) == 75
    assert level(percentage(maximum * .5, maximum)) == "medium"
    assert level(percentage(maximum * .3, maximum)) == "support"


def test_boundaries_blank_zero_and_no_rounding_promotion():
    assert level(percentage(None, 20)) == "missing"
    assert level(percentage(0, 20)) == "support"
    assert level(percentage(14.9999, 20)) == "medium"
    assert level(percentage(9.9999, 20)) == "support"
    for score, maximum in [(-1, 20), (21, 20), (float("nan"), 20), (1, 0), (1, float("inf"))]:
        with pytest.raises(ValueError):
            percentage(score, maximum)


def test_snapshot_matches_approved_example_and_filters():
    record = sample_record()
    result = build_snapshot(record, "ar")
    assert result["stats"] == {"total": 8, "graded": 7, "missing": 1, "mean": 58.57, "completion": 87.5}
    assert [x["count"] for x in result["distribution"]] == [3, 2, 2, 1]
    assert [x["mean"] for x in result["classes"]] == [61.25, 55]
    assert result["students"][0]["gap"] == 13.75
    assert result["students"][6]["percentage"] is None
    assert result["students"][7]["percentage"] == 0
    assert build_snapshot(record, "ar")["snapshot_id"] == result["snapshot_id"]
    assert build_snapshot(record, "en")["snapshot_id"] != result["snapshot_id"]
    filtered = build_snapshot(record, "ar", "c2")
    assert filtered["stats"]["mean"] == 55
    assert len(filtered["students"]) == 4
    record["scores"] = {}
    assert build_snapshot(record)["stats"]["mean"] is None


def test_owner_and_assigned_class_permissions(api_fixture):
    client, _, state = api_fixture
    base = "/api/baseline-assessments"
    params = {"school_section": "international", "academic_year": "2026-2027", "semester": 2, "quarter": 1}
    assert len(client.get(base, params=params).json()) == 1
    state["user"]["id"] = "another-teacher"
    assert client.get(base, params=params).json() == []
    assert client.get(base + "/sample").status_code == 404
    assert client.patch(base + "/sample/scores", json={"revision": 1, "scores": {"s0": 19}}).status_code == 404
    state["user"].update(id="t1", assigned_class_ids=["c1"])
    assert client.get(base, params=params).json() == []
    assert client.get(base + "/sample").status_code == 404
    state["user"]["role_name"] = "Admin"
    assert client.get(base + "/sample").status_code == 200
    state["user"]["role_name"] = "Counselor"
    assert client.get(base + "/sample").status_code == 403


def test_setup_scoping_and_immutable_maximum(api_fixture):
    client, db, state = api_fixture
    base = "/api/baseline-assessments"
    payload = {"title": "Computer science", "test_date": "2026-08-28", "max_score": 30,
               "school_section": "international", "academic_year": "2026-2027", "semester": 1, "quarter": 1, "class_ids": ["c1"]}
    response = client.post(base, json=payload)
    assert response.status_code == 201, response.text
    record = db.baseline_assessments.rows[-1]
    assert record["teacher_id"] == "t1" and len(record["roster"]) == 4
    assert record["scores"] == {} and record["max_score"] == 30
    assert client.patch(f"{base}/{record['id']}/scores", json={"revision": 1, "max_score": 20, "scores": {"s0": 10}}).status_code == 422
    for fields in [{"max_score": 0}, {"max_score": -1}, {"max_score": True}, {"title": "  "}, {"teacher_id": "other"}, {"class_ids": ["c1", "c1"]}]:
        assert client.post(base, json={**payload, **fields}).status_code == 422
    assert client.post(base, json={**payload, "class_ids": ["arabic"]}).status_code == 403
    state["user"]["role_name"] = "Admin"
    assert client.post(base, json={**payload, "class_ids": ["arabic"]}).status_code == 422
    assert client.post(base, json={**payload, "academic_year": "2027-2028"}).status_code == 422
    assert client.post(base, json={**payload, "school_section": "arabic", "class_ids": ["arabic"]}).status_code == 201
    assert len(db.baseline_assessments.rows[-1]["roster"]) == 1


def test_atomic_save_validation_and_concurrency(api_fixture):
    client, db, _ = api_fixture
    url = "/api/baseline-assessments/sample/scores"
    original = copy.deepcopy(db.baseline_assessments.rows[0])
    for changes in [{"s0": -1}, {"s0": 21}, {"outsider": 4}, {"s0": True}, {"s0": "NaN"}]:
        assert client.patch(url, json={"revision": 1, "scores": changes}).status_code == 422
    assert db.baseline_assessments.rows[0] == original
    assert client.patch(url, json={"revision": 1, "scores": {"s0": 0, "s1": None}}).status_code == 200
    saved = db.baseline_assessments.rows[0]
    assert saved["scores"]["s0"] == 0 and saved["scores"]["s1"] is None
    assert saved["scores"]["s2"] == original["scores"]["s2"]
    assert client.patch(url, json={"revision": 1, "scores": {"s0": 10}}).status_code == 409
    db.baseline_assessments.force_conflict = True
    assert client.patch(url, json={"revision": 2, "scores": {"s0": 10}}).status_code == 409


def test_metadata_rename_preserves_scope_roster_and_scores(api_fixture):
    client, db, state = api_fixture
    base = "/api/baseline-assessments/sample"
    original = copy.deepcopy(db.baseline_assessments.rows[0])
    payload = {
        "revision": 1,
        "title": "  اختبار تشخيصي محدث  ",
        "class_names": {"c1": "الصف الرابع أ", "c2": "الصف الرابع ب"},
    }
    response = client.patch(base + "/metadata", json=payload)
    assert response.status_code == 200, response.text
    assert response.json() == {
        "id": "sample",
        "title": "اختبار تشخيصي محدث",
        "classes": [
            {"id": "c1", "name": "الصف الرابع أ"},
            {"id": "c2", "name": "الصف الرابع ب"},
        ],
        "revision": 2,
    }
    saved = db.baseline_assessments.rows[0]
    assert saved["title"] == "اختبار تشخيصي محدث"
    assert saved["class_ids"] == original["class_ids"]
    assert saved["max_score"] == original["max_score"]
    assert saved["scores"] == original["scores"]
    assert [(row["id"], row["class_id"]) for row in saved["roster"]] == [
        (row["id"], row["class_id"]) for row in original["roster"]
    ]
    assert {row["class_name"] for row in saved["roster"] if row["class_id"] == "c1"} == {"الصف الرابع أ"}
    assert client.patch(base + "/metadata", json=payload).status_code == 409
    invalid = {"revision": 2, "title": "Updated", "class_names": {"c1": "4A"}}
    assert client.patch(base + "/metadata", json=invalid).status_code == 422
    state["user"].update(id="another-teacher", assigned_class_ids=["c1", "c2"])
    assert client.patch(base + "/metadata", json={**invalid, "class_names": {"c1": "4A", "c2": "4B"}}).status_code == 404


def test_list_scope_and_export_snapshot_guard(api_fixture):
    client, _, state = api_fixture
    base = "/api/baseline-assessments"
    params = {"school_section": "international", "academic_year": "2026-2027", "semester": 2, "quarter": 1}
    assert client.get(base, params={**params, "academic_year": "2026-2029"}).status_code == 422
    for changed in [{"school_section": "arabic"}, {"academic_year": "2027-2028"}, {"semester": 1}, {"quarter": 2}]:
        assert client.get(base, params={**params, **changed}).json() == []
    snap = client.get(base + "/sample", params={"lang": "ar", "class_id": "c1"}).json()
    export_params = {"lang": "ar", "class_id": "c1", "student_id": "s0", "snapshot_id": snap["snapshot_id"]}
    response = client.get(base + "/sample/export.pdf", params=export_params)
    assert response.status_code == 200 and response.content.startswith(b"%PDF")
    assert response.headers["x-baseline-snapshot"] == snap["snapshot_id"]
    assert client.get(base + "/sample/export.pdf", params={**export_params, "student_id": "s4"}).status_code == 422
    assert client.get(base + "/sample/export.pdf", params={**export_params, "lang": "en"}).status_code == 409
    client.patch(base + "/sample/scores", json={"revision": 1, "scores": {"s0": 16}})
    assert client.get(base + "/sample/export.pdf", params=export_params).status_code == 409
    state["user"]["id"] = "other"
    assert client.get(base + "/sample/export.pdf", params=export_params).status_code == 404


def test_delete_record_removes_only_the_revision_matched_snapshot(api_fixture):
    client, db, _ = api_fixture
    url = "/api/baseline-assessments/sample"
    stale = client.delete(url, params={"revision": 2})
    assert stale.status_code == 409
    assert len(db.baseline_assessments.rows) == 1
    db.baseline_assessments.rows[0].pop("title")  # Legacy records may predate a title.
    deleted = client.delete(url, params={"revision": 1})
    assert deleted.status_code == 200
    assert deleted.json()["title"] == "sample"
    assert deleted.json()["students_in_snapshot"] == 8
    assert db.baseline_assessments.rows == []


def test_excel_import_preview_apply_and_revision_guard(api_fixture):
    client, db, _ = api_fixture
    base = "/api/baseline-assessments/sample"
    content = score_file([("طالب تجريبي 01", 19), ("طالب غير موجود", 10), ("طالب تجريبي 02", 21)])
    preview = client.post(
        base + "/import",
        params={"revision": 1, "apply": "false", "class_id": "c1"},
        files={"file": ("scores.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["matched_count"] == 1
    assert preview.json()["unmatched_count"] == 1
    assert preview.json()["invalid_count"] == 1
    assert db.baseline_assessments.rows[0]["scores"]["s0"] == 15

    applied = client.post(
        base + "/import",
        params={"revision": 1, "apply": "true", "class_id": "c1"},
        files={"file": ("scores.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["imported_count"] == 1 and applied.json()["revision"] == 2
    assert db.baseline_assessments.rows[0]["scores"]["s0"] == 19
    stale = client.post(
        base + "/import",
        params={"revision": 1, "apply": "true"},
        files={"file": ("scores.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert stale.status_code == 409


def test_excel_export_uses_reference_template_and_snapshot_guard(api_fixture):
    client, _, _ = api_fixture
    base = "/api/baseline-assessments/sample"
    snapshot = client.get(base, params={"lang": "ar", "class_id": "c1"}).json()
    response = client.get(base + "/export.xlsx", params={"lang": "ar", "class_id": "c1", "snapshot_id": snapshot["snapshot_id"]})
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content)).active
    assert [cell.value for cell in sheet[1]] == REFERENCE_HEADERS
    assert sheet.max_row == 5
    assert sheet["A2"].value == "طالب تجريبي 01"
    assert sheet["D2"].value == 15
    assert sheet["B1"].fill.fgColor.rgb.endswith("D9D9D9")
    assert response.headers["x-baseline-snapshot"] == snapshot["snapshot_id"]
    assert client.get(base + "/export.xlsx", params={"lang": "en", "class_id": "c1", "snapshot_id": snapshot["snapshot_id"]}).status_code == 409


@pytest.mark.parametrize("lang", ["en", "ar"])
def test_pdf_contains_snapshot_values_text_and_embedded_arabic_fonts(lang):
    snapshot = build_snapshot(sample_record(), lang)
    pdf = render_baseline_pdf(snapshot, "s0")
    reader = PdfReader(io.BytesIO(pdf))
    assert len(reader.pages) == 3
    text = "\n".join(p.extract_text(extraction_mode="layout") for p in reader.pages)
    for expected in ["75%", "58.57%", "87.5%", "61.25%", "55%", "15 / 20", "2026-2027", "Q3", snapshot["snapshot_id"][:12]]:
        assert expected in text, expected
    assert shaped(snapshot["students"][0]["full_name"]) in text
    assert any("Amiri" in str(font.get_object().get("/BaseFont")) for page in reader.pages for font in page["/Resources"]["/Font"].get_object().values())


def test_large_pdf_preserves_every_student_and_empty_results():
    record = sample_record(100)
    snapshot = build_snapshot(record, "ar")
    reader = PdfReader(io.BytesIO(render_baseline_pdf(snapshot, "s99")))
    text = "\n".join(p.extract_text(extraction_mode="layout") for p in reader.pages)
    assert len(reader.pages) > 3
    assert all(shaped(s["full_name"]) in text for s in snapshot["students"])
    record["scores"] = {}
    empty = PdfReader(io.BytesIO(render_baseline_pdf(build_snapshot(record), "s0")))
    assert len(empty.pages) > 3
