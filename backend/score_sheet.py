"""Strict score-sheet parsing and reference-format Excel export helpers.

The uploaded school-platform sheet contains many metadata columns, but score
imports deliberately trust only the unshaded source columns needed for matching:
student name and submission grade. No enrollment or assessment metadata is read.
"""
from __future__ import annotations

import io
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Any, Callable, Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REFERENCE_HEADERS = [
    "اسم الطالب",
    "اسم المعلم",
    "المستوى",
    "درجة التسليم",
    "النسبة المئوية%",
    "تم التصحيح",
    "اسم الاختبار",
    "زمن الاختبار",
    "تم التسليم",
    "تاريخ النشر",
    "تاريخ التسليم",
]

NAME_HEADERS = {
    "اسمالطالب", "الطالب", "studentname", "student", "fullname", "studentfullname",
}
SCORE_HEADERS = {
    "درجةالتسليم", "درجهالتسليم", "submissiongrade", "submittedgrade", "submissionscore",
}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


class ScoreSheetError(ValueError):
    """Stable API-facing validation code."""


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[^\w\d]+", "", text, flags=re.UNICODE)


def normalize_student_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = re.sub(r"[\u064b-\u065f\u0670\u06d6-\u06ed\u0640]", "", text)
    text = text.translate(str.maketrans({"أ": "ا", "إ": "ا", "آ": "ا", "ٱ": "ا", "ى": "ي"}))
    return re.sub(r"[^\w\d]+", "", text, flags=re.UNICODE)


def parse_score_value(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, bool):
        raise ScoreSheetError("score_sheet_invalid_score")
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
        cleaned = cleaned.replace("٫", ".").replace("٬", "").replace(",", ".")
        try:
            number = float(cleaned)
        except ValueError as exc:
            raise ScoreSheetError("score_sheet_invalid_score") from exc
    else:
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ScoreSheetError("score_sheet_invalid_score") from exc
    if not math.isfinite(number):
        raise ScoreSheetError("score_sheet_invalid_score")
    return number


def _header_positions(values: Iterable[Any]) -> Dict[str, int]:
    positions: Dict[str, int] = {}
    for index, value in enumerate(values):
        normalized = normalize_header(value)
        if normalized in NAME_HEADERS and "name" not in positions:
            positions["name"] = index
        if normalized in SCORE_HEADERS and "score" not in positions:
            positions["score"] = index
    return positions


def read_score_sheet(content: bytes, filename: str) -> Dict[str, Any]:
    """Read only the exact student-name and submission-grade source columns."""
    lowered = (filename or "").lower()
    if not lowered.endswith((".xlsx", ".xls", ".csv")):
        raise ScoreSheetError("score_sheet_file_type")
    if not content or len(content) > MAX_UPLOAD_BYTES:
        raise ScoreSheetError("score_sheet_file_size")

    candidates: List[Tuple[str, pd.DataFrame]] = []
    try:
        if lowered.endswith(".csv"):
            candidates.append((filename or "CSV", pd.read_csv(io.BytesIO(content), header=None, dtype=object)))
        else:
            engine = "xlrd" if lowered.endswith(".xls") else None
            book = pd.ExcelFile(io.BytesIO(content), engine=engine)
            for sheet_name in book.sheet_names:
                candidates.append((sheet_name, book.parse(sheet_name, header=None, dtype=object)))
    except Exception as exc:
        raise ScoreSheetError("score_sheet_invalid_file") from exc

    selected = None
    for sheet_name, frame in candidates:
        for row_index in range(min(10, len(frame.index))):
            positions = _header_positions(frame.iloc[row_index].tolist())
            if "name" in positions and "score" in positions:
                selected = (sheet_name, frame, row_index, positions)
                break
        if selected:
            break
    if not selected:
        raise ScoreSheetError("score_sheet_missing_columns")

    sheet_name, frame, header_row, positions = selected
    rows: List[Dict[str, Any]] = []
    blank_name_count = 0
    for row_index in range(header_row + 1, len(frame.index)):
        row = frame.iloc[row_index]
        raw_name = row.iloc[positions["name"]] if positions["name"] < len(row) else None
        if raw_name is None or (isinstance(raw_name, float) and pd.isna(raw_name)) or not str(raw_name).strip():
            blank_name_count += 1
            continue
        raw_score = row.iloc[positions["score"]] if positions["score"] < len(row) else None
        rows.append({"row": row_index + 1, "name": str(raw_name).strip(), "raw_score": raw_score})
    if not rows:
        raise ScoreSheetError("score_sheet_no_students")
    return {"sheet_name": sheet_name, "rows": rows, "blank_name_count": blank_name_count}


def match_score_rows(
    imported_rows: List[Dict[str, Any]],
    roster: List[Dict[str, Any]],
    max_score_for: Callable[[Dict[str, Any]], float],
    current_score_for: Callable[[Dict[str, Any]], Any],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Match unique normalized names and validate against each target maximum."""
    roster_by_name: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for student in roster:
        key = normalize_student_name(student.get("full_name"))
        if key:
            roster_by_name[key].append(student)

    file_keys = [normalize_student_name(row["name"]) for row in imported_rows]
    duplicate_keys = {key for key, count in Counter(file_keys).items() if key and count > 1}
    matches: List[Dict[str, Any]] = []
    unmatched_names: List[str] = []
    ambiguous_names: List[str] = []
    invalid_rows: List[Dict[str, Any]] = []
    blank_scores = 0
    duplicate_count = 0
    new_count = overwrite_count = unchanged_count = 0

    for row, key in zip(imported_rows, file_keys):
        if key in duplicate_keys:
            duplicate_count += 1
            continue
        candidates = roster_by_name.get(key, [])
        if not candidates:
            unmatched_names.append(row["name"])
            continue
        if len(candidates) != 1:
            ambiguous_names.append(row["name"])
            continue
        student = candidates[0]
        try:
            score = parse_score_value(row.get("raw_score"))
            if score is None:
                blank_scores += 1
                continue
            maximum = float(max_score_for(student))
            if score < 0 or score > maximum:
                raise ScoreSheetError("score_sheet_score_out_of_range")
        except (ScoreSheetError, TypeError, ValueError) as exc:
            code = str(exc) if isinstance(exc, ScoreSheetError) else "score_sheet_invalid_score"
            invalid_rows.append({"row": row["row"], "name": row["name"], "reason": code})
            continue
        current = current_score_for(student)
        if current is None:
            new_count += 1
        elif float(current) == score:
            unchanged_count += 1
        else:
            overwrite_count += 1
        matches.append({"student_id": student["id"], "name": student.get("full_name"), "score": score, "current": current})

    summary = {
        "rows_total": len(imported_rows),
        "matched_count": len(matches),
        "new_count": new_count,
        "overwrite_count": overwrite_count,
        "unchanged_count": unchanged_count,
        "unmatched_count": len(unmatched_names),
        "ambiguous_count": len(ambiguous_names),
        "duplicate_count": duplicate_count,
        "invalid_count": len(invalid_rows),
        "blank_score_count": blank_scores,
        "unmatched_names": unmatched_names[:20],
        "ambiguous_names": ambiguous_names[:20],
        "invalid_rows": invalid_rows[:20],
    }
    return summary, matches


def _excel_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    return "" if value is None else value


def build_reference_score_workbook(rows: List[Dict[str, Any]]) -> bytes:
    """Export the exact 11-column reference order and shaded-column convention."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(1, len(rows) + 1)}"
    ws.append(REFERENCE_HEADERS)

    keys = [
        "student_name", "teacher_name", "level", "score", "percentage", "corrected",
        "test_name", "duration", "submitted_at", "published_at", "due_at",
    ]
    for row in rows:
        ws.append([_excel_value(row.get(key)) for key in keys])

    gray = PatternFill("solid", fgColor="D9D9D9")
    white = PatternFill("solid", fgColor="FFFFFF")
    shaded_columns = {2, 6, 7, 8, 9, 10, 11}
    for column in range(1, 12):
        fill = gray if column in shaded_columns else white
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=1, max_row=max(1, len(rows) + 1)):
            for item in cell:
                item.fill = fill
                item.font = Font(name="Calibri", size=11)
                item.alignment = Alignment(horizontal="right", vertical="center")
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=11, bold=True)

    widths = [28, 24, 14, 16, 18, 16, 34, 14, 22, 22, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row_index in range(2, len(rows) + 2):
        ws.cell(row_index, 5).number_format = "0.00"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
