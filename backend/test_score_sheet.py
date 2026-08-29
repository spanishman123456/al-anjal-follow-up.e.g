import asyncio
import io
from types import SimpleNamespace

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook

import server
from score_sheet import (
    REFERENCE_HEADERS,
    ScoreSheetError,
    build_reference_score_workbook,
    match_score_rows,
    read_score_sheet,
)


def score_file(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(REFERENCE_HEADERS)
    for name, score, level in rows:
        sheet.append([name, "بيانات مظللة لا تستورد", level, score, 0.5, "مظلل", "مظلل", 45, "مظلل", "مظلل", "مظلل"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_reads_only_student_and_submission_grade_and_matches_normalized_arabic():
    parsed = read_score_sheet(score_file([("أحمد علي", "٤٫٥", "4A")]), "scores.xlsx")
    assert parsed["rows"] == [{"row": 2, "name": "أحمد علي", "raw_score": "٤٫٥"}]
    summary, matches = match_score_rows(
        parsed["rows"],
        [{"id": "s1", "full_name": "احمد على", "quiz1": 2}],
        lambda _student: 5,
        lambda student: student.get("quiz1"),
    )
    assert summary["matched_count"] == 1
    assert summary["overwrite_count"] == 1
    assert matches[0]["score"] == 4.5


def test_duplicate_ambiguous_blank_and_out_of_range_rows_are_not_applied():
    rows = [
        {"row": 2, "name": "Duplicate", "raw_score": 4},
        {"row": 3, "name": "Duplicate", "raw_score": 3},
        {"row": 4, "name": "Same Name", "raw_score": 2},
        {"row": 5, "name": "Blank", "raw_score": None},
        {"row": 6, "name": "Too High", "raw_score": 6},
    ]
    roster = [
        {"id": "d", "full_name": "Duplicate"},
        {"id": "a1", "full_name": "Same Name"},
        {"id": "a2", "full_name": "Same Name"},
        {"id": "b", "full_name": "Blank"},
        {"id": "h", "full_name": "Too High"},
    ]
    summary, matches = match_score_rows(rows, roster, lambda _student: 5, lambda _student: None)
    assert matches == []
    assert summary["duplicate_count"] == 2
    assert summary["ambiguous_count"] == 1
    assert summary["blank_score_count"] == 1
    assert summary["invalid_count"] == 1
    assert summary["invalid_rows"][0]["reason"] == "score_sheet_score_out_of_range"


def test_reference_export_has_exact_columns_order_and_shading():
    content = build_reference_score_workbook([{
        "student_name": "طالب تجريبي",
        "teacher_name": "معلم تجريبي",
        "level": "4A",
        "score": 4,
        "percentage": 0.8,
        "corrected": "تم التصحيح",
        "test_name": "اختبار",
    }])
    sheet = load_workbook(io.BytesIO(content)).active
    assert [cell.value for cell in sheet[1]] == REFERENCE_HEADERS
    assert [sheet.cell(2, column).value for column in range(1, 6)] == ["طالب تجريبي", "معلم تجريبي", "4A", 4, 0.8]
    assert sheet["B1"].fill.fgColor.rgb.endswith("D9D9D9")
    assert sheet["F1"].fill.fgColor.rgb.endswith("D9D9D9")
    assert sheet["A1"].fill.fgColor.rgb.endswith("FFFFFF")
    assert sheet.sheet_view.rightToLeft is True


class FakeCollection:
    def __init__(self):
        self.operations = []

    async def find_one(self, _query, _projection=None):
        return {"id": "w1", "semester": 1, "quarter": 1, "number": 1}

    async def bulk_write(self, operations, ordered=False):
        assert ordered is False
        self.operations.extend(operations)


def test_international_route_writes_only_selected_quiz_and_rejects_practical(monkeypatch):
    weeks = FakeCollection()
    scores = FakeCollection()
    monkeypatch.setattr(server, "db", SimpleNamespace(weeks=weeks, student_scores=scores))

    async def fake_students(**_kwargs):
        return [{"id": "s1", "full_name": "Test Student", "quiz1": 1, "chapter_test1_practical": 9}]

    async def fake_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(server, "get_students", fake_students)
    monkeypatch.setattr(server, "log_user_action", fake_log)
    content = score_file([("Test Student", 4, "4A")])
    user = {"id": "admin", "role_name": "Admin"}

    preview = asyncio.run(server.import_score_sheet(
        file=UploadFile(filename="scores.xlsx", file=io.BytesIO(content)),
        context="international_quiz", target="quiz1", apply=False, week_id="w1",
        academic_year="2026-2027", semester=1, quarter=1, class_id=None, current_user=user,
    ))
    assert preview["matched_count"] == 1 and scores.operations == []

    result = asyncio.run(server.import_score_sheet(
        file=UploadFile(filename="scores.xlsx", file=io.BytesIO(content)),
        context="international_quiz", target="quiz1", apply=True, week_id="w1",
        academic_year="2026-2027", semester=1, quarter=1, class_id=None, current_user=user,
    ))
    assert result["imported_count"] == 1
    update = scores.operations[0]._doc["$set"]
    assert update["quiz1"] == 4
    assert "quiz2" not in update and "chapter_test1_practical" not in update

    with pytest.raises(HTTPException) as error:
        asyncio.run(server.import_score_sheet(
            file=UploadFile(filename="scores.xlsx", file=io.BytesIO(content)),
            context="international_quiz", target="chapter_test1_practical", apply=True, week_id="w1",
            academic_year="2026-2027", semester=1, quarter=1, class_id=None, current_user=user,
        ))
    assert error.value.status_code == 422
    assert error.value.detail == "score_sheet_invalid_target"


def test_arabic_route_uses_student_raw_max_and_never_writes_practical(monkeypatch):
    arabic_scores = FakeCollection()
    monkeypatch.setattr(server, "db", SimpleNamespace(arabic_quarter_scores=arabic_scores))

    async def fake_payload(*_args, **_kwargs):
        return {"students": [{
            "id": "s1", "full_name": "طالب عربي", "exam_raw_max": 20,
            "theory_test_1": 10, "theory_test_2": None, "practical_test": 19,
        }]}

    async def fake_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(server, "build_arabic_grading_payload", fake_payload)
    monkeypatch.setattr(server, "log_user_action", fake_log)
    content = score_file([("طالب عربي", 18, "السادس")])
    result = asyncio.run(server.import_score_sheet(
        file=UploadFile(filename="scores.xlsx", file=io.BytesIO(content)),
        context="arabic_theory", target="theory_test_2", apply=True, week_id=None,
        academic_year="2026-2027", semester=1, quarter=1, class_id=None,
        current_user={"id": "admin", "role_name": "Admin"},
    ))
    assert result["imported_count"] == 1
    update = arabic_scores.operations[0]._doc["$set"]
    assert update["theory_test_2"] == 18
    assert "theory_test_1" not in update and "practical_test" not in update

    with pytest.raises(HTTPException) as error:
        asyncio.run(server.import_score_sheet(
            file=UploadFile(filename="scores.xlsx", file=io.BytesIO(content)),
            context="arabic_theory", target="practical_test", apply=True, week_id=None,
            academic_year="2026-2027", semester=1, quarter=1, class_id=None,
            current_user={"id": "admin", "role_name": "Admin"},
        ))
    assert error.value.detail == "score_sheet_invalid_target"


def test_missing_required_headers_has_stable_error():
    workbook = Workbook()
    workbook.active.append(["اسم الطالب", "عمود آخر"])
    output = io.BytesIO()
    workbook.save(output)
    with pytest.raises(ScoreSheetError, match="score_sheet_missing_columns"):
        read_score_sheet(output.getvalue(), "bad.xlsx")
